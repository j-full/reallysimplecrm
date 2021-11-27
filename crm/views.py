from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from accounts.mixins import UserIsOwnerMixin
from django.utils import timezone
from django.contrib import messages
from django.views.generic import ListView, DetailView, UpdateView, DeleteView

from openpyxl import Workbook, load_workbook
from .forms import ContactForm
from .models import Contact, PostCard

class DashboardView(LoginRequiredMixin, ListView):
    Model = Contact
    template_name = "contacts_list.html"
    context_object_name = 'contacts'

    def get_queryset(self):
        queryset = Contact.objects.filter(created_by=self.request.user).order_by('last_name', 'first_name')
        return queryset
    

@login_required
def contact_new(request):
    if request.method == 'POST':
        form = ContactForm(request.POST, request.FILES)
        if form.is_valid():
            new_contact = form.save(commit=False)
            new_contact.created_by = request.user
            new_contact.save()
            messages.success(request,'New Contact Added')
            return redirect(reverse_lazy('contact_detail', kwargs={'pk':new_contact.pk}))
    else:
        form = ContactForm()
    return render(request, 'contact_form.html', {'form': form, 'form_title': 'New Contact'})



class ContactEdit(LoginRequiredMixin, UserIsOwnerMixin, UpdateView):
    model = Contact
    form_class = ContactForm
    template_name = 'contact_form.html'

    def get_context_data(self, **kwargs):
        context = super(ContactEdit, self).get_context_data(**kwargs)
        context['form_title'] = f'Edit Contact: {self.object.first_name} {self.object.last_name}'
        return context
    
    def get_success_url(self):
        return reverse_lazy('contact_detail', kwargs={'pk': self.kwargs['pk']})


class ContactDetail(LoginRequiredMixin, UserIsOwnerMixin, DetailView):
    model = Contact
    template_name = "contact_detail.html"

    def get_context_data(self, **kwargs):
        context = super(ContactDetail, self).get_context_data(**kwargs)
        context['can_send_postcard'] = self.object.can_send_postcard
        context.update(dict((field.name, getattr(self.object, field.name)) for field in self.object._meta.fields))
        context['postcards'] = PostCard.objects.filter(contact=self.object).order_by('-time_sent')
        return context


class ContactDelete(LoginRequiredMixin, UserIsOwnerMixin, DeleteView):
    model = Contact
    template_name = 'contact_confirm_delete.html'
    success_url = reverse_lazy('dashboard')

@login_required
def send_postcard(request, *args, **kwargs):
    contact = get_object_or_404(Contact, pk=kwargs['pk'], created_by=request.user)
    can_send = contact.can_send_postcard
    if request.method == 'POST' and can_send:
        postcard = PostCard(contact=contact, time_sent=timezone.now())
        postcard.save()
        messages.success(request,'Postcard Send!')
        return redirect(reverse_lazy('contact_detail', kwargs={'pk':contact.pk}))
    return render(request, 'confirm_send_postcard.html', {'object':contact, 'can_send':can_send})

@login_required
def export_xls(request):
    contacts = Contact.objects.filter(created_by=request.user)
    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Contacts"
    sheet['A1'] = 'First Name'
    sheet['B1'] = 'Last Name'
    sheet['C1'] = 'Address 1'
    sheet['D1'] = 'Address 2'
    sheet['E1'] = 'City'
    sheet['F1'] = 'State'
    sheet['G1'] = 'Zip Code'

    for index, contact in enumerate(contacts, start=2):
        sheet[f'A{index}'] = contact.first_name
        sheet[f'B{index}'] = contact.last_name
        sheet[f'C{index}'] = contact.address1
        sheet[f'D{index}'] = contact.address2
        sheet[f'E{index}'] = contact.city
        sheet[f'F{index}'] = contact.state
        sheet[f'G{index}'] = contact.zip_code

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=ReallySimpleCRMExport.xlsx'
    workbook.save(response)
    return response

@login_required
def import_xls(request):
    if request.method == 'POST':
        workbook = load_workbook(request.FILES['xls'])
        sheet = workbook.active
        new_contacts = []
        for row in sheet.iter_rows(min_row=2):
            contact = Contact(
                first_name=row[0].value,
                last_name=row[1].value,
                address1=str(row[2].value or ''),
                address2=str(row[3].value or ''),
                city=str(row[4].value or ''),
                state=str(row[5].value or ''),
                zip_code=str(row[6].value or ''),
                created_by=request.user
            )
            new_contacts.append(contact)
        Contact.objects.bulk_create(new_contacts)
        messages.success(request,'Excel Sheet Imported!')
        return redirect(reverse_lazy('dashboard'))
    return render(request, 'import_xls.html')